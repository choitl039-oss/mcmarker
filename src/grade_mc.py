import argparse
import os


# Placeholder modules


def _get_poppler_path():
    """Return a Poppler bin directory if one is configured locally.

    Streamlit Cloud installs Poppler utilities into PATH, so we should not
    force a Windows-specific absolute path there.
    """

    for env_name in ("POPPLER_PATH", "POPLER_PATH"):
        candidate = os.environ.get(env_name)
        if candidate and os.path.isdir(candidate):
            return candidate

    if os.name == "nt":
        for candidate in (
            r"C:\Users\dmshw00039\poppler-25.12.0\Library\bin",
            r"C:\Program Files\poppler\Library\bin",
            r"C:\Program Files\poppler\bin",
        ):
            if os.path.isdir(candidate):
                return candidate

    return None

def convert_pdf_to_images(pdf_path):
    # Use pdf2image to convert each page to an image
    from pdf2image import convert_from_path

    if not os.path.exists(pdf_path):
        raise FileNotFoundError(f"PDF file not found: {pdf_path}")

    # returns a list of PIL.Image objects
    poppler_path = _get_poppler_path()
    if poppler_path:
        images = convert_from_path(pdf_path, poppler_path=poppler_path)
    else:
        images = convert_from_path(pdf_path)
    return images


def detect_answers(image, layout=None):
    """Return detected answers and fields from a scanned image.

    The returned tuple is ``(answers, fields)`` where ``answers`` is a
    dict mapping question numbers to choice letters (A/B/C/D) and
    ``fields`` is a dict for additional bubbled information such as
    class or class number.  The layout may contain special entries whose
    keys start with ``field_``; these are treated as separate groups of
    bubbles rather than standard questions.  For example:

        "field_class": {"A": [...], "B": [...], ...}
        "field_classnum1": {"0": [...], "1": [...], ...}
        "field_classnum2": {"0": [...], ...}

    The detector processes both question and field regions using the
    same dark/saturation heuristic.  Fields are collapsed into strings
    (class letter or two-digit number) for easy reporting.
    """

    import cv2
    import numpy as np

    # convert PIL image to numpy array (RGB->BGR for OpenCV)
    img = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    if layout:
        answers = {}
        fields = {}
        for key, choices in layout.items():
            is_field = key.startswith("field_")
            best_choice = None
            best_val = -1  # start low, pick the HIGHEST percentage
            print(f"\nProcessing {key}:")
            for choice, box in choices.items():
                x, y, w, h = box
                # Expand ROI to 45x45 and center it on (x,y)
                roi_size = 45
                x_start = max(0, x - roi_size // 2)
                y_start = max(0, y - roi_size // 2)
                x_end = min(img.shape[1], x_start + roi_size)
                y_end = min(img.shape[0], y_start + roi_size)
                roi = img[y_start : y_end, x_start : x_end]
                if roi.size == 0:
                    print(f"  Choice {choice}: empty ROI {box}")
                    continue
                # convert ROI to HSV to detect colored/dark marks
                hsv = cv2.cvtColor(roi, cv2.COLOR_BGR2HSV)
                h_chan, s_chan, v_chan = cv2.split(hsv)
                # count pixels that are either dark or sufficiently saturated (colored)
                dark_mask = v_chan < 200
                color_mask = s_chan > 50
                mask = dark_mask | color_mask
                marked = cv2.countNonZero(mask.astype('uint8'))
                total = mask.size
                pct = marked/total*100
                print(f"  Choice {choice}: marked pct = {pct:.1f}% ({marked}/{total})")
                # save roi for inspection
                debug_root = os.path.join(os.getcwd(), "debug_rois")
                os.makedirs(debug_root, exist_ok=True)
                cv2.imwrite(os.path.join(debug_root, f"{key}_{choice}.png"), roi)
                if pct > best_val:  # choose region with HIGHEST marked percentage
                    best_val = pct
                    best_choice = choice
            if is_field:
                fields[key] = best_choice
                print(f"  → Field {key} = {best_choice}")
            else:
                answers[key] = best_choice
                print(f"  → Detected: {best_choice}")
        print(f"Debug images saved in ./debug_rois")
        return answers, fields

    # fallback: threshold and find all dark blobs, then assign sequentially
    thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
    contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    boxes = [cv2.boundingRect(c) for c in contours]
    boxes = sorted(boxes, key=lambda b: (b[1], b[0]))
    detected = {}
    for i, (x, y, w, h) in enumerate(boxes, start=1):
        question = str(i)
        detected[question] = chr(64 + ((i - 1) % 4) + 1)
    return detected, {}


def grade_sheet(detected, answer_key):
    # Compare detected answers against key
    # detected is expected to be dict {question: choice}
    results = {}
    for q, ans in answer_key.items():
        results[q] = (detected.get(q) == ans)
    return results


def generate_report(results, output_path, answer_key):
    """Generate styled Excel report.

    Columns order:
    student, form_class, class_number, Q1..Qn, total_score, percentage

    Each question cell shows the detected MC option. Green means correct,
    red means incorrect.
    """

    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    question_numbers = sorted(answer_key.keys(), key=lambda x: int(x))
    total_questions = len(question_numbers)

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    headers = ["student", "form_class", "class_number"]
    headers += [f"Q{q}" for q in question_numbers]
    headers += ["total_score", "percentage"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for student, rec in results.items():
        detected = rec.get("detected", {})
        form_class = rec.get("form_class", "")
        class_number = rec.get("class_number", "")

        row = [student, form_class, class_number]
        correctness = []
        score = 0

        for q in question_numbers:
            picked = detected.get(q) or ""
            row.append(picked)
            is_correct = (picked == answer_key.get(q))
            correctness.append(is_correct)
            if is_correct:
                score += 1

        pct = (score / total_questions * 100) if total_questions else 0
        row.append(f"{score}/{total_questions}")
        row.append(f"{pct:.0f}%")
        ws.append(row)

        row_idx = ws.max_row
        # keep class_number as text (preserve leading zero)
        ws.cell(row=row_idx, column=3).number_format = "@"
        first_q_col = 4
        for i, is_correct in enumerate(correctness):
            cell = ws.cell(row=row_idx, column=first_q_col + i)
            if cell.value:
                cell.fill = green_fill if is_correct else red_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # simple column sizing
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    for col in range(4, 4 + len(question_numbers)):
        ws.column_dimensions[get_column_letter(col)].width = 6
    ws.column_dimensions[get_column_letter(4 + len(question_numbers))].width = 12
    ws.column_dimensions[get_column_letter(5 + len(question_numbers))].width = 12

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Grade MC answer sheets from scanned PDF")
    parser.add_argument("--input", required=True, help="Path to scanned PDF containing answer sheets")
    parser.add_argument("--key", required=True, help="Path to JSON file with answer key")
    parser.add_argument("--output", required=True, help="Path to Excel report to generate")
    parser.add_argument("--layout", help="Optional JSON file describing bubble locations on the template")
    args = parser.parse_args()

    import json

    # load answer key
    if not os.path.exists(args.key):
        raise FileNotFoundError(f"Answer key file not found: {args.key}")
    with open(args.key, "r") as f:
        answer_key = json.load(f)

    images = convert_pdf_to_images(args.input)

    # optionally load layout description
    layout = None
    if args.layout:
        if not os.path.exists(args.layout):
            raise FileNotFoundError(f"Layout file not found: {args.layout}")
        with open(args.layout, "r") as lf:
            layout = json.load(lf)

    # In a real implementation we'd identify student IDs and iterate
    results = {}
    for idx, img in enumerate(images, start=1):
        detected, fields = detect_answers(img, layout=layout)

        # combine form + class into one cell
        form_val = fields.pop('field_form', '')
        class_val = fields.pop('field_class', '')
        if form_val and class_val:
            form_class = f"{form_val}-{class_val}"
        else:
            form_class = form_val or class_val

        # two-digit class number (e.g. 0 + 3 => "03")
        d1 = fields.pop('field_classnum1', None)
        d2 = fields.pop('field_classnum2', None)
        class_number = ""
        if d1 is not None and d2 is not None:
            class_number = f"{d1}{d2}"

        # use page number as placeholder student id
        results[f"page_{idx}"] = {
            "detected": detected,
            "form_class": form_class,
            "class_number": class_number,
            "extra_fields": fields,
        }

    generate_report(results, args.output, answer_key)
    print(f"Report written to {args.output}")


if __name__ == "__main__":
    main()
