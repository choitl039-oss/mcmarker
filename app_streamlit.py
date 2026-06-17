import json
import os
import tempfile
import io
from pathlib import Path

import cv2
import numpy as np
import pandas as pd
import streamlit as st
from PIL import Image

from src.grade_mc import convert_pdf_to_images, detect_answers, generate_report


ROOT = Path(__file__).resolve().parent
LAYOUT_PATH = ROOT / "layout.json"
ANSWER_KEY_PATH = ROOT / "answer_key.json"
TEMPLATE_PDF_PATH = ROOT / "MC-blank.pdf"


def list_layout_files(root: Path) -> list[Path]:
    files = sorted(root.glob("*.json"))
    return [p for p in files if p.name.lower().startswith("layout")]


def list_template_pdf_files(root: Path) -> list[Path]:
    return sorted(root.glob("*.pdf"))


def _normalize_profile_token(stem: str) -> str:
    token = stem.lower()
    for prefix in ["layout", "mc-blank", "mc_blank", "template", "blank", "form"]:
        token = token.replace(prefix, "")
    token = token.replace("__", "_").replace("--", "-")
    token = token.strip("_- ")
    return token or "default"


def build_template_profiles(layout_files: list[Path], template_files: list[Path]) -> dict:
    """Create profile map {profile_name: {layout: Path, template: Path}}."""
    layout_map = {}
    template_map = {}

    for p in layout_files:
        layout_map[_normalize_profile_token(p.stem)] = p
    for p in template_files:
        template_map[_normalize_profile_token(p.stem)] = p

    profiles = {}
    common_tokens = sorted(set(layout_map.keys()) & set(template_map.keys()))
    for token in common_tokens:
        display = token.replace("_", " ").replace("-", " ").title()
        profiles[display] = {"layout": layout_map[token], "template": template_map[token]}

    # Fallback: if nothing matches by token, pair defaults by original constants.
    if not profiles and layout_files and template_files:
        default_layout = next((p for p in layout_files if p.name == LAYOUT_PATH.name), layout_files[0])
        default_template = next((p for p in template_files if p.name == TEMPLATE_PDF_PATH.name), template_files[0])
        profiles["Default"] = {"layout": default_layout, "template": default_template}

    return profiles


def load_json(path: Path):
    if not path.exists():
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def build_answer_key_template_csv(question_keys: list[str]) -> bytes:
    lines = ["question,answer"]
    for q in sorted(question_keys, key=lambda x: int(x)):
        lines.append(f"{q},")
    return "\n".join(lines).encode("utf-8")


def parse_uploaded_answer_key(answer_key_file) -> dict:
    """Parse uploaded answer key from JSON, CSV, or XLSX.

    Supported tabular schema:
    - question,answer
    """
    name = (answer_key_file.name or "").lower()

    if name.endswith(".json"):
        loaded = json.load(answer_key_file)
        parsed = {}
        for q, a in loaded.items():
            ans = str(a).strip().upper()
            if ans in ["A", "B", "C", "D"]:
                parsed[str(int(q))] = ans
        return parsed

    if name.endswith(".csv"):
        df = pd.read_csv(io.BytesIO(answer_key_file.getvalue()))
    elif name.endswith(".xlsx"):
        df = pd.read_excel(io.BytesIO(answer_key_file.getvalue()))
    else:
        raise ValueError("Unsupported file type. Use JSON, CSV, or XLSX.")

    cols = {c.lower().strip(): c for c in df.columns}
    if "question" not in cols or "answer" not in cols:
        raise ValueError("Template columns must be: question,answer")

    q_col = cols["question"]
    a_col = cols["answer"]

    parsed = {}
    for _, row in df.iterrows():
        q_val = row.get(q_col)
        a_val = row.get(a_col)
        if pd.isna(q_val) or pd.isna(a_val):
            continue

        try:
            q_key = str(int(q_val))
        except Exception:
            continue

        ans = str(a_val).strip().upper()
        if ans in ["A", "B", "C", "D"]:
            parsed[q_key] = ans

    return parsed


def apply_page_offset(layout: dict, page_num: int) -> dict:
    """
    Apply page-specific Y offsets to layout coordinates.
    
    Supports both:
    1. Uniform page offset (e.g., "6-7": -24)
    2. Block-specific offsets within a page (e.g., "12__1-5": -10, "12__6-10": -15)
    
    Checks session state for custom offsets, falls back to defaults.
    """
    # Check if user has defined custom page offsets in session state
    custom_offsets = st.session_state.get("custom_page_offsets", {})
    page_offset_ranges = st.session_state.get("page_offset_ranges", {})
    
    # Check for block-specific offsets first (format: "pagenum__1-5", "pagenum__6-10")
    block_offsets = {}
    for range_key, offset in page_offset_ranges.items():
        if "__" in range_key:
            page_str, block_range = range_key.split("__")
            if int(page_str) == page_num:
                # This is a block-specific offset for this page
                block_offsets[block_range] = offset
    
    if block_offsets:
        # Apply block-specific offsets
        adjusted_layout = {}
        for key, choices in layout.items():
            if key.startswith("field_"):
                # Don't offset field coordinates
                adjusted_layout[key] = choices
            else:
                # Check which block this question belongs to
                try:
                    q_num = int(key)
                    question_offset = 0
                    
                    # Find the matching block
                    for block_range, offset in block_offsets.items():
                        start_q, end_q = map(int, block_range.split("-"))
                        if start_q <= q_num <= end_q:
                            question_offset = offset
                            break
                    
                    # Apply offset
                    adjusted_layout[key] = {}
                    for choice, (x, y, w, h) in choices.items():
                        adjusted_layout[key][choice] = [x, y + question_offset, w, h]
                except:
                    adjusted_layout[key] = choices
        
        return adjusted_layout
    
    # Otherwise use uniform page offset
    y_offset = 0
    
    # First check if there's a custom offset for this specific page
    if str(page_num) in custom_offsets:
        y_offset = custom_offsets[str(page_num)]
    else:
        # Then check page ranges
        for range_key, offset in page_offset_ranges.items():
            if "__" not in range_key:  # Skip block-specific offsets
                try:
                    start, end = map(int, range_key.split("-"))
                    if start <= page_num <= end:
                        y_offset = offset
                        break
                except:
                    pass
        
        # Default offsets if no custom ones set
        if not page_offset_ranges:
            page_offsets = {
                (1, 5): 0,      # pages 1-5: no offset
                (6, 7): -24,    # pages 6-7: shift up 24 pixels
                (8, 11): 0,     # pages 8-11: no offset
                (12, 22): 0,    # pages 12-22: no offset (for now)
            }
            
            for (start, end), offset in page_offsets.items():
                if start <= page_num <= end:
                    y_offset = offset
                    break
    
    if y_offset == 0:
        return layout
    
    # Apply offset to all question coordinates
    adjusted_layout = {}
    for key, choices in layout.items():
        if key.startswith("field_"):
            # Don't offset field coordinates (they're absolute positions)
            adjusted_layout[key] = choices
        else:
            # Apply y_offset to question coordinates
            adjusted_layout[key] = {}
            for choice, (x, y, w, h) in choices.items():
                adjusted_layout[key][choice] = [x, y + y_offset, w, h]
    
    return adjusted_layout


def detect_page_offset_from_question_numbers(image, layout: dict) -> dict:
    """
    Automatically detect Y-offset by finding printed question numbers.
    
    Returns a dict mapping question blocks to their offsets:
    {
        "1-5": offset1,
        "6-10": offset2,
        "11-15": offset3,
        ...
    }
    
    This allows different parts of the same page to have different offsets
    (useful when scanner shifts the paper during scanning).
    """
    import cv2
    import numpy as np
    import pytesseract

    # Set Tesseract path (default Windows installation)
    tesseract_default_path = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
    if os.path.exists(tesseract_default_path):
        pytesseract.pytesseract.tesseract_cmd = tesseract_default_path
    
    img = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    def _detect_block_offsets_by_bubble_edges(gray_img: np.ndarray, layout_data: dict) -> dict:
        """
        Fallback alignment that does not rely on OCR.
        For each 5-question block, search Y-shift that maximizes edge density
        around expected bubble ROIs.
        """
        edges = cv2.Canny(gray_img, 80, 180)
        blocks_local = [
            (1, 5), (6, 10), (11, 15), (16, 20), (21, 25),
            (26, 30), (31, 35), (36, 40)
        ]

        def score_shift_for_block(start_q: int, end_q: int, shift: int) -> int:
            score = 0
            for q_num in range(start_q, end_q + 1):
                q_key = str(q_num)
                if q_key not in layout_data:
                    continue
                for _, (x, y, w, h) in layout_data[q_key].items():
                    cy = y + shift
                    x1 = max(0, int(x - w // 2 - 6))
                    y1 = max(0, int(cy - h // 2 - 6))
                    x2 = min(edges.shape[1], int(x + w // 2 + 6))
                    y2 = min(edges.shape[0], int(cy + h // 2 + 6))
                    if x2 > x1 and y2 > y1:
                        roi = edges[y1:y2, x1:x2]
                        score += int(cv2.countNonZero(roi))
            return score

        result = {}
        for start_q, end_q in blocks_local:
            # coarse search
            best_shift = 0
            best_score = -1
            for shift in range(-80, 81, 4):
                s = score_shift_for_block(start_q, end_q, shift)
                if s > best_score:
                    best_score = s
                    best_shift = shift

            # fine search around the coarse optimum
            fine_best_shift = best_shift
            fine_best_score = best_score
            for shift in range(best_shift - 4, best_shift + 5):
                s = score_shift_for_block(start_q, end_q, shift)
                if s > fine_best_score:
                    fine_best_score = s
                    fine_best_shift = shift

            result[f"{start_q}-{end_q}"] = int(fine_best_shift)

        return result
    
    # Expected Y positions for all questions
    expected_positions = {}
    for q_num in range(1, 41):  # Check Q1-Q40
        q_key = str(q_num)
        if q_key in layout:
            # Get Y position from the layout (use choice A as reference)
            expected_y = layout[q_key]["A"][1]
            expected_positions[q_num] = expected_y
    
    if not expected_positions:
        return {"1-40": 0}
    
    # Define search region (left side where question numbers are printed)
    search_region = gray[0:gray.shape[0], 50:400]
    
    # Use OCR to find all text in the region
    detected_positions = {}
    block_detection_counts = {
        "1-5": 0,
        "6-10": 0,
        "11-15": 0,
        "16-20": 0,
        "21-25": 0,
        "26-30": 0,
        "31-35": 0,
        "36-40": 0,
    }
    try:
        # Configure tesseract to only look for digits
        custom_config = r'--oem 3 --psm 6 outputbase digits'
        ocr_data = pytesseract.image_to_data(search_region, config=custom_config, output_type=pytesseract.Output.DICT)
        
        # Parse OCR results
        for i, text in enumerate(ocr_data['text']):
            text_clean = text.strip()
            if text_clean.isdigit() and 1 <= int(text_clean) <= 40:
                q_num = int(text_clean)
                y_pos = ocr_data['top'][i]
                conf_raw = ocr_data['conf'][i]
                try:
                    conf = float(conf_raw)
                except Exception:
                    conf = -1
                
                # Only use high-confidence detections
                if conf > 40:
                    if q_num not in detected_positions:
                        detected_positions[q_num] = y_pos
        
        # Debug: Return detected positions for inspection
        print(f"DEBUG OCR: Found {len(detected_positions)} question numbers: {sorted(detected_positions.keys())}")
        
        # Calculate offsets for each 5-question block
        blocks = [
            (1, 5), (6, 10), (11, 15), (16, 20), (21, 25),
            (26, 30), (31, 35), (36, 40)
        ]
        
        block_offsets = {}
        
        for start_q, end_q in blocks:
            # Find detected questions in this block
            block_offsets_list = []
            for q_num in range(start_q, end_q + 1):
                if q_num in detected_positions and q_num in expected_positions:
                    detected_y = detected_positions[q_num]
                    expected_y = expected_positions[q_num]
                    offset = detected_y - expected_y
                    block_offsets_list.append(offset)
            
            block_key = f"{start_q}-{end_q}"
            block_detection_counts[block_key] = len(block_offsets_list)

            if block_offsets_list:
                # Use median offset for this block
                median_offset = int(np.median(block_offsets_list))
                block_offsets[block_key] = median_offset
                print(f"  Block Q{start_q}-{end_q}: offset={median_offset}px from {len(block_offsets_list)} detections")
            else:
                # No detection in this block, use 0
                block_offsets[block_key] = 0

        # Fallback: for weak/no OCR blocks, use bubble-edge alignment
        edge_offsets = _detect_block_offsets_by_bubble_edges(gray, layout)
        for block_key in block_offsets.keys():
            # Require at least 2 OCR anchors in the block; otherwise fallback.
            if block_detection_counts.get(block_key, 0) < 2:
                block_offsets[block_key] = edge_offsets.get(block_key, block_offsets[block_key])

        return block_offsets if block_offsets else edge_offsets
    
    except Exception as e:
        print(f"OCR detection failed: {e}")
        # Full fallback when OCR fails completely
        return _detect_block_offsets_by_bubble_edges(gray, layout)


def detect_page_offset_from_fields(image, layout: dict) -> int:
    """
    Automatically detect Y-offset by comparing field positions.
    
    Uses the form/class fields as reference points since they have
    fixed positions on the page. Returns the calculated Y-offset.
    """
    import cv2
    import numpy as np
    
    img = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)
    
    # Try to find a clear reference field (field_form digit 5 is usually marked)
    reference_fields = ["field_form", "field_class"]
    
    detected_y_positions = []
    expected_y_positions = []
    
    for field_key in reference_fields:
        if field_key not in layout:
            continue
        
        field_choices = layout[field_key]
        
        # For each choice, find bright/dark regions
        for choice, (x, y_expected, w, h) in field_choices.items():
            roi_size = 45
            x_start = max(0, x - roi_size // 2)
            y_start = max(0, y_expected - roi_size // 2)
            x_end = min(img.shape[1], x_start + roi_size)
            y_end = min(img.shape[0], y_start + roi_size)
            
            roi = img[y_start:y_end, x_start:x_end]
            if roi.size == 0:
                continue
            
            # Check if this bubble is marked
            hsv = cv2.cvtColor(roi, cv2.COLOR_BGR2HSV)
            h_chan, s_chan, v_chan = cv2.split(hsv)
            dark_mask = v_chan < 200
            color_mask = s_chan > 50
            mask = dark_mask | color_mask
            marked = cv2.countNonZero(mask.astype('uint8'))
            total = mask.size
            pct = marked / total * 100
            
            # If this bubble is strongly marked (>15%), it's a reference point
            if pct > 15:
                # Now scan vertically to find the actual center of the mark
                # Use a sliding window approach
                best_y_offset = 0
                best_marked_pct = pct
                
                # Try offsets from -50 to +50 to find where mark is strongest
                for test_offset in range(-50, 51, 2):
                    test_y_start = max(0, y_expected + test_offset - roi_size // 2)
                    test_y_end = min(img.shape[0], test_y_start + roi_size)
                    test_roi = img[test_y_start:test_y_end, x_start:x_end]
                    
                    if test_roi.size == 0:
                        continue
                    
                    test_hsv = cv2.cvtColor(test_roi, cv2.COLOR_BGR2HSV)
                    _, test_s, test_v = cv2.split(test_hsv)
                    test_mask = (test_v < 200) | (test_s > 50)
                    test_marked = cv2.countNonZero(test_mask.astype('uint8'))
                    test_pct = test_marked / test_mask.size * 100
                    
                    if test_pct > best_marked_pct:
                        best_marked_pct = test_pct
                        best_y_offset = test_offset
                
                detected_y_positions.append(y_expected + best_y_offset)
                expected_y_positions.append(y_expected)
    
    if len(detected_y_positions) > 0:
        # Calculate median offset
        offsets = [detected - expected for detected, expected in zip(detected_y_positions, expected_y_positions)]
        offset = int(np.median(offsets))
        return offset
    
    return 0


def get_question_keys_from_layout(layout: dict):
    return sorted([k for k in layout.keys() if k.isdigit()], key=lambda x: int(x))


def draw_mapping_overlay(pil_image, layout: dict, show_fields: bool = True):
    rgb = np.array(pil_image.convert("RGB"))
    bgr = cv2.cvtColor(rgb, cv2.COLOR_RGB2BGR)

    for key, choices in layout.items():
        is_field = key.startswith("field_")
        if is_field and not show_fields:
            continue

        color = (0, 255, 255) if is_field else (0, 255, 0)
        for choice, box in choices.items():
            x, y, w, h = box

            # draw exactly the ROI used by detect_answers
            roi_size = 45
            x1 = max(0, int(x - roi_size // 2))
            y1 = max(0, int(y - roi_size // 2))
            x2 = min(bgr.shape[1] - 1, x1 + roi_size)
            y2 = min(bgr.shape[0] - 1, y1 + roi_size)

            cv2.rectangle(bgr, (x1, y1), (x2, y2), color, 1)
            label = f"{key}:{choice}"
            cv2.putText(
                bgr,
                label,
                (x1, max(15, y1 - 2)),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.30,
                color,
                1,
                cv2.LINE_AA,
            )

    return cv2.cvtColor(bgr, cv2.COLOR_BGR2RGB)


def _find_triangle_anchors(pil_image):
    """Find 4 registration triangles near page corners.

    Returns dict with keys: tl, tr, bl, br -> (x, y) or None if not found.
    """
    bgr = cv2.cvtColor(np.array(pil_image.convert("RGB")), cv2.COLOR_RGB2BGR)
    gray = cv2.cvtColor(bgr, cv2.COLOR_BGR2GRAY)
    h, w = gray.shape[:2]

    blur = cv2.GaussianBlur(gray, (5, 5), 0)
    bw = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]

    contours, _ = cv2.findContours(bw, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    candidates = []
    min_area = max(80, int(h * w * 0.00002))
    max_area = int(h * w * 0.01)

    for c in contours:
        area = cv2.contourArea(c)
        if area < min_area or area > max_area:
            continue

        peri = cv2.arcLength(c, True)
        if peri <= 0:
            continue

        approx = cv2.approxPolyDP(c, 0.04 * peri, True)
        if len(approx) != 3:
            continue

        x, y, ww, hh = cv2.boundingRect(approx)
        if ww < 8 or hh < 8:
            continue

        m = cv2.moments(approx)
        if m["m00"] == 0:
            continue
        cx = float(m["m10"] / m["m00"])
        cy = float(m["m01"] / m["m00"])

        # Keep triangles close to page borders (where registration marks usually are)
        edge_margin_x = int(0.22 * w)
        edge_margin_y = int(0.22 * h)
        near_left = cx < edge_margin_x
        near_right = cx > (w - edge_margin_x)
        near_top = cy < edge_margin_y
        near_bottom = cy > (h - edge_margin_y)
        if not ((near_left or near_right) and (near_top or near_bottom)):
            continue

        candidates.append((cx, cy))

    if len(candidates) < 4:
        return None

    corner_targets = {
        "tl": (0.0, 0.0),
        "tr": (float(w - 1), 0.0),
        "bl": (0.0, float(h - 1)),
        "br": (float(w - 1), float(h - 1)),
    }

    anchors = {}
    remaining = candidates.copy()
    for key, (tx, ty) in corner_targets.items():
        if not remaining:
            return None
        best_idx = min(
            range(len(remaining)),
            key=lambda i: (remaining[i][0] - tx) ** 2 + (remaining[i][1] - ty) ** 2,
        )
        anchors[key] = remaining.pop(best_idx)

    return anchors


def align_page_with_triangle_anchors(page_image, template_image):
    """Align scanned page to template using 4 triangle anchors.

    Returns (aligned_image, success, message).
    """
    if page_image is None or template_image is None:
        return page_image, False, "Template/page not available"

    template_anchors = _find_triangle_anchors(template_image)
    page_anchors = _find_triangle_anchors(page_image)

    if template_anchors is None:
        return page_image, False, "Template triangle anchors not found"
    if page_anchors is None:
        return page_image, False, "Page triangle anchors not found"

    src = np.float32([
        page_anchors["tl"],
        page_anchors["tr"],
        page_anchors["br"],
        page_anchors["bl"],
    ])
    dst = np.float32([
        template_anchors["tl"],
        template_anchors["tr"],
        template_anchors["br"],
        template_anchors["bl"],
    ])

    h_t, w_t = np.array(template_image).shape[:2]
    page_bgr = cv2.cvtColor(np.array(page_image.convert("RGB")), cv2.COLOR_RGB2BGR)
    matrix = cv2.getPerspectiveTransform(src, dst)
    warped = cv2.warpPerspective(page_bgr, matrix, (w_t, h_t), flags=cv2.INTER_LINEAR, borderMode=cv2.BORDER_REPLICATE)
    aligned = Image.fromarray(cv2.cvtColor(warped, cv2.COLOR_BGR2RGB))
    return aligned, True, "Aligned by triangle anchors"


def build_results_for_export(review_records, answer_key_subset):
    results = {}
    for rec in review_records:
        student_id = rec["student_id"]
        detected = rec["final_detected"]
        fields = rec.get("fields", {}).copy()

        form_val = fields.pop("field_form", "")
        class_val = fields.pop("field_class", "")
        form_class = f"{form_val}-{class_val}" if form_val and class_val else (form_val or class_val)

        d1 = fields.pop("field_classnum1", "")
        d2 = fields.pop("field_classnum2", "")
        class_number = f"{d1}{d2}" if (d1 != "" or d2 != "") else ""

        results[student_id] = {
            "detected": detected,
            "form_class": form_class,
            "class_number": class_number,
            "extra_fields": fields,
        }

    return results


def build_report_bytes(results, answer_key_subset):
    """Return the styled Excel report as bytes for direct browser download."""

    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    question_numbers = sorted(answer_key_subset.keys(), key=lambda x: int(x))
    total_questions = len(question_numbers)

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    headers = ["student", "form_class", "class_number"]
    headers += [f"Q{q}" for q in question_numbers]
    headers += ["total_score", "percentage"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for student, rec in results.items():
        detected = rec.get("detected", {})
        form_class = rec.get("form_class", "")
        class_number = rec.get("class_number", "")

        row = [student, form_class, class_number]
        correctness = []
        score = 0

        for q in question_numbers:
            picked = detected.get(q) or ""
            row.append(picked)
            is_correct = (picked == answer_key_subset.get(q))
            correctness.append(is_correct)
            if is_correct:
                score += 1

        pct = (score / total_questions * 100) if total_questions else 0
        row.append(f"{score}/{total_questions}")
        row.append(f"{pct:.0f}%")
        ws.append(row)

        row_idx = ws.max_row
        ws.cell(row=row_idx, column=3).number_format = "@"
        first_q_col = 4
        for i, is_correct in enumerate(correctness):
            cell = ws.cell(row=row_idx, column=first_q_col + i)
            if cell.value:
                cell.fill = green_fill if is_correct else red_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    for col in range(4, 4 + len(question_numbers)):
        ws.column_dimensions[get_column_letter(col)].width = 6
    ws.column_dimensions[get_column_letter(4 + len(question_numbers))].width = 12
    ws.column_dimensions[get_column_letter(5 + len(question_numbers))].width = 12

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def main():
    st.set_page_config(page_title="MC Marker App", layout="wide")
    st.title("MC Marker App")
    st.caption("Upload scanned PDFs, review each page, edit answers if needed, and export styled Excel.")

    layout_candidates = list_layout_files(ROOT)
    template_candidates = list_template_pdf_files(ROOT)
    template_profiles = build_template_profiles(layout_candidates, template_candidates)

    with st.sidebar:
        st.header("Settings")

        st.subheader("Template Selection")
        if not layout_candidates:
            st.error("No layout JSON found. Add a file like layout.json")
            return
        if not template_candidates:
            st.error("No template PDF found in project folder")
            return

        selected_layout_path = None
        selected_template_path = None

        if template_profiles:
            profile_names = list(template_profiles.keys())
            selected_profile = st.selectbox(
                "Template Profile",
                options=profile_names,
                index=0,
                help="Choose paper type. Layout + template are selected automatically.",
            )
            selected_layout_path = template_profiles[selected_profile]["layout"]
            selected_template_path = template_profiles[selected_profile]["template"]
        else:
            st.warning("No auto-matched profiles found. Using manual selection.")

            default_layout_idx = 0
            for i, p in enumerate(layout_candidates):
                if p.name == LAYOUT_PATH.name:
                    default_layout_idx = i
                    break

            default_template_idx = 0
            for i, p in enumerate(template_candidates):
                if p.name == TEMPLATE_PDF_PATH.name:
                    default_template_idx = i
                    break

            selected_layout_name = st.selectbox(
                "Layout JSON",
                options=[p.name for p in layout_candidates],
                index=default_layout_idx,
                help="Choose the coordinate mapping for the paper type.",
            )
            selected_template_name = st.selectbox(
                "Template PDF",
                options=[p.name for p in template_candidates],
                index=default_template_idx,
                help="Choose the blank form PDF for preview/alignment.",
            )
            selected_layout_path = ROOT / selected_layout_name
            selected_template_path = ROOT / selected_template_name

        layout = load_json(selected_layout_path)
        if not layout:
            st.error(f"Layout file not found or empty: {selected_layout_path}")
            return

        question_keys = get_question_keys_from_layout(layout)
        if not question_keys:
            st.error(f"No numeric question keys found in: {selected_layout_path.name}")
            return

        max_questions = len(question_keys)

        st.caption(f"Using layout: {selected_layout_path.name}")
        st.caption(f"Using template: {selected_template_path.name}")

        question_count = st.number_input(
            "How many questions to grade",
            min_value=1,
            max_value=max_questions,
            value=min(40, max_questions),
            step=1,
        )
        use_triangle_alignment = st.checkbox("Use triangle-anchor alignment", value=True)
        show_fields_overlay = st.checkbox("Show form/class fields on template overlay", value=True)

        st.divider()
        st.subheader("Answer Key")

        template_bytes = build_answer_key_template_csv(question_keys)
        st.download_button(
            "⬇️ Download answer key template (CSV)",
            data=template_bytes,
            file_name="answer_key_template.csv",
            mime="text/csv",
            help="Fill the answer column with A/B/C/D, then upload it back.",
        )
        st.caption("Template format: question,answer")
        
        answer_key_mode = st.radio("Select answer key source:", ["Use file", "Enter manually"], horizontal=True)
        
        if answer_key_mode == "Use file":
            answer_key_file = st.file_uploader(
                "Upload answer key file (JSON / CSV / XLSX)",
                type=["json", "csv", "xlsx"],
                key="answer_key_uploader",
            )
            
            if answer_key_file is not None:
                try:
                    answer_key = parse_uploaded_answer_key(answer_key_file)
                    st.success(f"Loaded answer key: {len(answer_key)} questions")
                except Exception as e:
                    st.error(f"Failed to load answer key: {e}")
                    answer_key = load_json(ANSWER_KEY_PATH)
            else:
                answer_key = load_json(ANSWER_KEY_PATH)
                if answer_key:
                    st.info(f"Using default answer_key.json: {len(answer_key)} questions")
        else:
            st.write("Enter answers as comma-separated letters (e.g., A,B,C,D,A,B,C,D,...)")
            answers_input = st.text_area("Answers:", placeholder="A,B,C,D,A,B,C,D,A,B,C,D,A,B,C,D,A,B,C,D,A,B,C,D,A", height=80)
            
            if answers_input.strip():
                try:
                    answers_list = [a.strip().upper() for a in answers_input.split(",")]
                    answer_key = {str(i+1): ans for i, ans in enumerate(answers_list) if ans in ["A", "B", "C", "D"]}
                    st.success(f"Parsed {len(answer_key)} answers")
                except Exception as e:
                    st.error(f"Failed to parse answers: {e}")
                    answer_key = {}
            else:
                answer_key = {}

        st.divider()
        output_name = st.text_input("Output Excel filename", value="mc_report.xlsx")
        if not output_name.lower().endswith(".xlsx"):
            output_name += ".xlsx"

    template_img = None
    if selected_template_path.exists():
        try:
            template_img = convert_pdf_to_images(str(selected_template_path))[0]
        except Exception as e:
            st.warning(f"Template loading failed: {e}")

    tabs = st.tabs(["1) Template & Mapping", "2) Calibrate Page Offsets", "3) Upload & Detect", "4) Review & Export"])

    # --- Tab 1: template mapping preview ---
    with tabs[0]:
        st.subheader("Blank template with mapping overlay")
        if template_img is not None:
            overlay = draw_mapping_overlay(template_img, layout, show_fields=show_fields_overlay)
            st.image(overlay, caption="Green = questions, Yellow = form/class fields", width='stretch')
        else:
            st.warning(f"Template PDF not found: {selected_template_path}")

    # --- Tab 2: calibrate page offsets ---
    with tabs[1]:
        st.subheader("Calibrate Page-Specific Y-Offsets")
        
        calibration_mode = st.radio("Choose calibration mode:", ["Auto-detect from PDF", "Single Page", "Batch (Manual Entry)"], horizontal=True)
        
        if calibration_mode == "Auto-detect from PDF":
            st.info("🤖 Automatically detect offsets using printed question numbers (1-40) as reference points. Detects offsets for each 5-question block (Q1-5, Q6-10, etc.) to handle scanner shifts.")
            
            auto_pdf = st.file_uploader("Upload PDF to auto-detect offsets", type=["pdf"], key="auto_pdf")
            
            if auto_pdf is not None:
                if st.button("🔍 Auto-detect offsets for all pages", type="primary"):
                    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                        tmp.write(auto_pdf.read())
                        auto_pdf_path = tmp.name
                    
                    try:
                        auto_images = convert_pdf_to_images(auto_pdf_path)
                        st.write(f"Analyzing {len(auto_images)} pages using OCR on question numbers (Q1-40)...")
                        
                        progress_bar = st.progress(0)
                        detected_offsets = {}  # {page_num: {block_range: offset}}
                        
                        for i, page_img in enumerate(auto_images, start=1):
                            if use_triangle_alignment and template_img is not None:
                                page_img, _, _ = align_page_with_triangle_anchors(page_img, template_img)
                            # Get block-specific offsets for this page
                            st.write(f"**Processing page {i}...**")
                            block_offsets = detect_page_offset_from_question_numbers(page_img, layout)
                            st.write(f"  Detected: {block_offsets}")
                            detected_offsets[i] = block_offsets
                            progress_bar.progress(i / len(auto_images))
                        
                        st.success("✓ Auto-detection complete!")
                        
                        # Display results in expandable sections per page
                        for page_num in sorted(detected_offsets.keys()):
                            block_offsets = detected_offsets[page_num]
                            
                            # Check if all blocks have same offset (uniform page)
                            unique_offsets = set(block_offsets.values())
                            if len(unique_offsets) == 1:
                                uniform_offset = list(unique_offsets)[0]
                                st.write(f"**Page {page_num}**: {uniform_offset:+d} px (uniform)")
                            else:
                                with st.expander(f"**Page {page_num}**: Multiple blocks detected"):
                                    for block_range, offset in sorted(block_offsets.items(), key=lambda x: int(x[0].split('-')[0])):
                                        st.write(f"  Q{block_range}: {offset:+d} px")
                        
                        # Build suggested ranges
                        st.write("---")
                        st.write("**Applying detected offsets:**")
                        
                        suggested_ranges = {}
                        for page_num, block_offsets in detected_offsets.items():
                            # Check if uniform or block-specific
                            unique_offsets = set(block_offsets.values())
                            
                            if len(unique_offsets) == 1:
                                # Uniform offset - use page range format
                                uniform_offset = list(unique_offsets)[0]
                                range_key = f"{page_num}-{page_num}"
                                suggested_ranges[range_key] = uniform_offset
                            else:
                                # Block-specific offsets - use page__block format
                                for block_range, offset in block_offsets.items():
                                    range_key = f"{page_num}__{block_range}"
                                    suggested_ranges[range_key] = offset
                        
                        st.write(f"Found {len(suggested_ranges)} offset entries")
                        
                        if st.button("💾 Apply detected offsets", type="primary"):
                            st.session_state["page_offset_ranges"] = suggested_ranges
                            st.success("✓ Offsets applied! Block-specific adjustments will be used during detection.")
                    
                    finally:
                        os.unlink(auto_pdf_path)
        
        elif calibration_mode == "Batch (Manual Entry)":
            st.info("📋 Define offsets for multiple page ranges at once, then apply to all pages during detection.")
            
            with st.form("batch_calibration_form"):
                st.write("**Define page ranges and their offsets:**")
                
                num_ranges = st.number_input("Number of page ranges", min_value=1, max_value=10, value=3)
                
                batch_ranges = {}
                for i in range(num_ranges):
                    col1, col2, col3 = st.columns([1, 1, 1])
                    with col1:
                        start_page = st.number_input(f"Range {i+1} Start", min_value=1, value=i*7 + 1, key=f"batch_start_{i}")
                    with col2:
                        end_page = st.number_input(f"Range {i+1} End", min_value=1, value=(i+1)*7, key=f"batch_end_{i}")
                    with col3:
                        offset = st.number_input(f"Offset (px)", value=0, key=f"batch_offset_{i}")
                    
                    if start_page <= end_page:
                        batch_ranges[f"{start_page}-{end_page}"] = offset
                
                submitted = st.form_submit_button("💾 Save Batch Configuration", type="primary")
                
                if submitted:
                    st.session_state["page_offset_ranges"] = batch_ranges
                    st.success("✓ Batch configuration saved! It will be applied when you run detection.")
                    st.write("**Current configuration:**")
                    for range_key, offset in batch_ranges.items():
                        st.write(f"  Pages {range_key}: **{offset:+d}** pixels")
        
        else:  # Single page mode
            st.info("🔍 Fine-tune offset for a single page with live preview.")
            
            # Upload PDF for calibration
            cal_pdf = st.file_uploader("Upload PDF for calibration", type=["pdf"], key="cal_pdf")
            
            if cal_pdf is not None:
                with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                    tmp.write(cal_pdf.read())
                    cal_pdf_path = tmp.name
                
                try:
                    cal_images = convert_pdf_to_images(cal_pdf_path)
                    
                    col1, col2 = st.columns([2, 1])
                    with col1:
                        page_num = st.slider("Select page to calibrate", 1, len(cal_images), 6)
                    with col2:
                        st.write(f"**Total pages: {len(cal_images)}**")
                    
                    cal_img = cal_images[page_num - 1]
                    if use_triangle_alignment and template_img is not None:
                        cal_img, aligned_ok, aligned_msg = align_page_with_triangle_anchors(cal_img, template_img)
                        if aligned_ok:
                            st.caption("Triangle-anchor alignment applied before calibration.")
                        else:
                            st.caption(f"Triangle-anchor alignment skipped: {aligned_msg}")
                    cal_mode = st.radio("Calibration mode:", ["Manual", "Auto-calibrate"], horizontal=True, key="single_mode")
                    
                    if cal_mode == "Auto-calibrate":
                        if st.button("🔍 Find best offset automatically", type="primary"):
                            st.write("Testing offsets -50 to +50...")
                            progress_bar = st.progress(0)
                            results_list = []
                            
                            for test_offset in range(-50, 51, 5):
                                preview_layout = apply_page_offset(layout, page_num)
                                # Apply manual offset
                                preview_layout = {
                                    k: {
                                        c: [x, y + test_offset, w, h]
                                        for c, (x, y, w, h) in v.items()
                                    }
                                    if not k.startswith("field_") else v
                                    for k, v in preview_layout.items()
                                }
                                
                                detected, fields = detect_answers(cal_img, layout=preview_layout)
                                
                                # Count high-confidence detections (Q1, Q5, Q10)
                                confidence_score = 0
                                for q_key in ["1", "5", "10"]:
                                    if q_key in detected and detected[q_key]:
                                        confidence_score += 1
                                
                                results_list.append({
                                    "offset": test_offset,
                                    "Q1": detected.get("1", "?"),
                                    "Q5": detected.get("5", "?"),
                                    "Q10": detected.get("10", "?"),
                                    "score": confidence_score
                                })
                                
                                progress_bar.progress((test_offset + 50) / 100)
                            
                            # Find best
                            best = max(results_list, key=lambda x: x["score"])
                            st.success(f"✓ Best offset found: **{best['offset']}** pixels")
                            
                            results_df = pd.DataFrame(results_list)
                            st.dataframe(results_df, use_container_width=True)
                            
                            if st.button(f"Use offset {best['offset']}", type="primary", key="use_auto"):
                                st.session_state[f"page_offset_{page_num}"] = best["offset"]
                                st.success(f"✓ Saved offset {best['offset']} for page {page_num}")
                    
                    else:  # Manual mode
                        offset_value = st.slider(
                            "Y-axis offset (pixels)",
                            min_value=-100,
                            max_value=100,
                            value=0,
                            step=1,
                            help="Negative = shift up, Positive = shift down",
                            key="manual_offset_slider"
                        )
                        
                        # Preview with offset
                        if st.button("Preview with offset", type="primary"):
                            preview_layout = apply_page_offset(layout, page_num)
                            # Apply manual offset
                            preview_layout = {
                                k: {
                                    c: [x, y + offset_value, w, h]
                                    for c, (x, y, w, h) in v.items()
                                }
                                if not k.startswith("field_") else v
                                for k, v in preview_layout.items()
                            }
                            
                            detected, fields = detect_answers(cal_img, layout=preview_layout)
                            
                            # Create image with ROI highlights
                            import cv2
                            preview_img = cv2.cvtColor(np.array(cal_img.convert("RGB")), cv2.COLOR_RGB2BGR)
                            
                            # Highlight Q1, Q5, Q10 bubbles
                            highlight_questions = ["1", "5", "10"]
                            for q_key in highlight_questions:
                                if q_key in preview_layout:
                                    choices = preview_layout[q_key]
                                    # Draw box around A choice (largest detected)
                                    x, y, w, h = choices["A"]
                                    roi_size = 45
                                    x1 = max(0, int(x - roi_size // 2))
                                    y1 = max(0, int(y - roi_size // 2))
                                    x2 = min(preview_img.shape[1] - 1, x1 + roi_size)
                                    y2 = min(preview_img.shape[0] - 1, y1 + roi_size)
                                    
                                    color = (0, 255, 0)  # green
                                    cv2.rectangle(preview_img, (x1, y1), (x2, y2), color, 3)
                                    cv2.putText(
                                        preview_img,
                                        f"Q{q_key}:{detected.get(q_key, '?')}",
                                        (x1, max(15, y1 - 5)),
                                        cv2.FONT_HERSHEY_SIMPLEX,
                                        0.8,
                                        color,
                                        2,
                                    )
                            
                            st.image(cv2.cvtColor(preview_img, cv2.COLOR_BGR2RGB), caption=f"Page {page_num} with offset {offset_value}px", use_container_width=True)
                            
                            col1, col2, col3 = st.columns(3)
                            with col1:
                                st.metric("Q1 detected", detected.get("1", "?"))
                            with col2:
                                st.metric("Q5 detected", detected.get("5", "?"))
                            with col3:
                                st.metric("Q10 detected", detected.get("10", "?"))
                            
                            if st.button(f"✓ Save offset {offset_value} for page {page_num}", type="primary", key="save_manual"):
                                st.session_state[f"page_offset_{page_num}"] = offset_value
                                st.success(f"✓ Saved offset {offset_value} for page {page_num}")
                
                finally:
                    os.unlink(cal_pdf_path)

    # --- Tab 3: upload and detect ---
    with tabs[2]:
        st.subheader("Upload answer-sheet PDF")
        uploaded_pdf = st.file_uploader("Choose PDF", type=["pdf"])

        if uploaded_pdf is not None and st.button("Run detection", type="primary"):
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                tmp.write(uploaded_pdf.read())
                tmp_pdf_path = tmp.name

            try:
                images = convert_pdf_to_images(tmp_pdf_path)
                selected_questions = question_keys[:question_count]

                review_records = []
                for i, page_img in enumerate(images, start=1):
                    if use_triangle_alignment and template_img is not None:
                        page_img, _, _ = align_page_with_triangle_anchors(page_img, template_img)
                    # Apply page-specific Y offsets
                    page_layout = apply_page_offset(layout, i)
                    detected, fields = detect_answers(page_img, layout=page_layout)
                    final_detected = {q: detected.get(q, "") for q in selected_questions}

                    review_records.append(
                        {
                            "student_id": f"page_{i}",
                            "page_index": i,
                            "image": np.array(page_img.convert("RGB")),
                            "fields": fields,
                            "final_detected": final_detected,
                        }
                    )

                st.session_state["review_records"] = review_records
                st.session_state["selected_questions"] = selected_questions
                st.success(f"Detection complete for {len(review_records)} page(s). Go to 'Review & Export'.")
            finally:
                try:
                    os.remove(tmp_pdf_path)
                except OSError:
                    pass

    # --- Tab 4: review and export ---
    with tabs[3]:
        st.subheader("Review each page before export")

        review_records = st.session_state.get("review_records", [])
        selected_questions = st.session_state.get("selected_questions", question_keys[:question_count])

        if not review_records:
            st.info("No detection results yet. Please upload a PDF and run detection first.")
            return

        for idx, rec in enumerate(review_records):
            st.markdown(f"### {rec['student_id']}")
            col1, col2 = st.columns([1.4, 1])

            with col1:
                st.image(rec["image"], caption=f"Preview of {rec['student_id']}", width='stretch')

            with col2:
                fields = rec.get("fields", {})
                form_val = fields.get("field_form", "")
                class_val = fields.get("field_class", "")
                d1 = fields.get("field_classnum1", "")
                d2 = fields.get("field_classnum2", "")
                st.write("Detected fields")
                st.write(f"- Form: {form_val}")
                st.write(f"- Class: {class_val}")
                st.write(f"- Class number: {d1}{d2}")

            editor_df = pd.DataFrame(
                {
                    "question": selected_questions,
                    "detected": [rec["final_detected"].get(q, "") for q in selected_questions],
                    "final_choice": [rec["final_detected"].get(q, "") for q in selected_questions],
                }
            )

            edited = st.data_editor(
                editor_df,
                key=f"editor_{idx}",
                width='stretch',
                hide_index=True,
                column_config={
                    "question": st.column_config.TextColumn(disabled=True),
                    "detected": st.column_config.TextColumn(disabled=True),
                    "final_choice": st.column_config.SelectboxColumn(options=["", "A", "B", "C", "D"]),
                },
            )

            rec["final_detected"] = {
                str(q): str(a)
                for q, a in zip(edited["question"].tolist(), edited["final_choice"].tolist())
            }

            st.divider()

        st.session_state["review_records"] = review_records

        missing_key = [q for q in selected_questions if q not in answer_key]
        if missing_key:
            st.warning(
                f"answer_key.json is missing these questions: {', '.join(missing_key)}. "
                "They will be excluded from scoring."
            )

        answer_key_subset = {q: answer_key[q] for q in selected_questions if q in answer_key}

        if st.button("Export Excel", type="primary"):
            if not answer_key_subset:
                st.error("No valid questions from answer_key.json for current question count.")
            else:
                results = build_results_for_export(review_records, answer_key_subset)
                out_path = ROOT / output_name
                generate_report(results, str(out_path), answer_key_subset)
                st.success(f"Exported: {out_path}")
                excel_bytes = build_report_bytes(results, answer_key_subset)
                st.download_button(
                    "Download Excel file",
                    data=excel_bytes,
                    file_name=output_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )


if __name__ == "__main__":
    main()
